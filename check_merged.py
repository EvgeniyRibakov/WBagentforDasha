"""Проверка объединённых ячеек."""
from openpyxl import load_workbook

file_path = 'data/11.12.2025/beautylab_11.12.2025.xlsx'
wb = load_workbook(file_path)
ws = wb.active

print("Объединённые ячейки в первой строке:")
print(ws.merged_cells.ranges)

print("\nФормат первой строки (первые 5 ячеек):")
for i in range(1, 6):
    cell = ws.cell(1, i)
    print(f"  {chr(64+i)}: value={cell.value}, merged={cell.coordinate in ws.merged_cells}")

