"""Скрипт для исправления заголовков в файлах отчётов (версия 2)."""
from pathlib import Path
from openpyxl import load_workbook
import time

# Правильные заголовки по столбцам A-P
CORRECT_HEADERS = [
    "Бренд",           # A (1)
    "Предмет",         # B (2)
    "Сезон",           # C (3)
    "Коллекция",       # D (4)
    "Наименование",    # E (5)
    "Артикул поставщика",  # F (6)
    "Номенклатура",    # G (7)
    "Баркод",          # H (8)
    "Размер",          # I (9)
    "Контракт",        # J (10)
    "Склад",           # K (11)
    "Заказано шт",     # L (12)
    "Заказано себестоимость",  # M (13)
    "Выкупили шт",     # N (14)
    "Выкупили руб",    # O (15)
    "Текущий остаток"  # P (16)
]

def fix_file_headers(file_path: Path):
    """Исправляет заголовки в файле."""
    print(f"\n{'='*60}")
    print(f"Обработка: {file_path.name}")
    print('='*60)
    
    # Загружаем файл
    wb = load_workbook(file_path)
    ws = wb.active
    
    # Показываем текущую первую строку
    print("ТЕКУЩАЯ первая строка:")
    for i in range(1, 17):
        value = ws.cell(1, i).value
        print(f"  {chr(64+i)} (col {i}): {value}")
    
    # Удаляем первую строку
    print("\n>>> Удаление первой строки...")
    ws.delete_rows(1)
    
    # Записываем правильные заголовки
    print("\n>>> Запись новых заголовков...")
    for col_idx, header in enumerate(CORRECT_HEADERS, start=1):
        ws.cell(row=1, column=col_idx).value = header
        print(f"  {chr(64+col_idx)} (col {col_idx}): {header}")
    
    # Сохраняем с явным указанием пути
    print("\n>>> Сохранение файла...")
    wb.save(str(file_path.absolute()))
    wb.close()
    
    # Небольшая задержка
    time.sleep(0.5)
    
    # Проверяем результат
    print("\n>>> Проверка результата...")
    wb_check = load_workbook(file_path)
    ws_check = wb_check.active
    
    print("НОВАЯ первая строка:")
    success = True
    for i in range(1, 17):
        value = ws_check.cell(1, i).value
        expected = CORRECT_HEADERS[i-1]
        match = "✓" if value == expected else "✗"
        print(f"  {chr(64+i)} (col {i}): {value} {match}")
        if value != expected:
            success = False
    
    wb_check.close()
    
    if success:
        print("\n✅ Файл успешно обработан!")
    else:
        print("\n⚠️ Обнаружены расхождения!")
    
    return success

def main():
    """Обрабатывает все файлы в папке data/11.12.2025."""
    data_dir = Path("data/11.12.2025")
    
    if not data_dir.exists():
        print(f"Папка {data_dir} не найдена!")
        return
    
    # Получаем все xlsx файлы (исключая временные ~$)
    xlsx_files = [f for f in data_dir.glob("*.xlsx") if not f.name.startswith("~$")]
    
    if not xlsx_files:
        print(f"Файлы .xlsx не найдены в {data_dir}")
        return
    
    print(f"\n{'='*60}")
    print(f"Найдено файлов: {len(xlsx_files)}")
    print('='*60)
    
    results = {}
    for file_path in xlsx_files:
        try:
            success = fix_file_headers(file_path)
            results[file_path.name] = success
        except Exception as e:
            print(f"\n✗ Ошибка: {e}")
            results[file_path.name] = False
    
    # Итоговый отчёт
    print(f"\n{'='*60}")
    print("ИТОГОВЫЙ ОТЧЁТ")
    print('='*60)
    for filename, success in results.items():
        status = "✅ OK" if success else "❌ ОШИБКА"
        print(f"{status}: {filename}")
    
    print(f"\n{'='*60}")
    print("✅ Обработка завершена!")
    print('='*60)

if __name__ == "__main__":
    main()

