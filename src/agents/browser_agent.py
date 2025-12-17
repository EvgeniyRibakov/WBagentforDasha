"""Агент для автоматизации работы с браузером Wildberries."""
import os
import re
import time
import shutil
import subprocess
from datetime import datetime, timedelta, date
from pathlib import Path
from typing import Optional, Dict, List

import undetected_chromedriver as uc
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, NoSuchElementException, StaleElementReferenceException
from openpyxl import load_workbook
from loguru import logger

from src.config.settings import Settings


class BrowserAgent:
    """Агент для автоматизации работы с браузером Wildberries."""

    # URL страницы отчётов Wildberries (явно указан в коде)
    WILDBERRIES_REPORTS_URL = "https://seller.wildberries.ru/analytics-reports/sales"

    # Список кабинетов для обработки
    CABINETS: List[Dict[str, str]] = [
        {"name": "MAU", "id": "53607"},
        {"name": "MAB", "id": "121614"},
        {"name": "MMA", "id": "174711"},
        {"name": "cosmo", "id": "224650"},
        {"name": "dreamlab", "id": "1140223"},
        {"name": "beautylab", "id": "4428365"},
    ]

    def __init__(self, settings: Settings):
        """Инициализация агента.

        Args:
            settings: Настройки приложения
        """
        self.settings = settings
        self.driver: Optional[uc.Chrome] = None
        self.downloads_dir = Path(settings.downloads_dir).resolve()
        self.data_dir = Path(settings.data_dir).resolve()
        self.example_first_stroke_path = Path(settings.example_first_stroke_file).resolve()

        # Создаём необходимые папки
        self.downloads_dir.mkdir(parents=True, exist_ok=True)
        self.data_dir.mkdir(parents=True, exist_ok=True)

    def start_browser(self) -> None:
        """Запуск Yandex Browser."""
        try:
            self._start_yandex_browser()
        except Exception as e:
            logger.error(f"Ошибка при запуске Yandex Browser: {e}")
            raise


    def _start_yandex_browser(self) -> None:
        """Запуск Yandex Browser с профилем из настроек (.env)."""
        options = uc.ChromeOptions()

        # Определяем путь к Yandex Browser
        if self.settings.yandex_browser_path:
            browser_path = Path(os.path.expandvars(self.settings.yandex_browser_path)).expanduser()
        else:
            browser_path = Path(os.path.expandvars("%LOCALAPPDATA%")) / "Yandex" / "YandexBrowser" / "Application" / "browser.exe"
        
        if not browser_path.exists():
            logger.error(f"✗ Yandex Browser не найден: {browser_path}")
            raise Exception(f"Yandex Browser не найден: {browser_path}")

        options.binary_location = str(browser_path.absolute())
        logger.info(f"✓ Yandex Browser: {browser_path}")

        # Используем изолированный профиль для сохранения авторизации
        automation_profile = Path("./yandex_automation_profile").resolve()
        automation_profile.mkdir(parents=True, exist_ok=True)
        
        options.add_argument(f'--user-data-dir={str(automation_profile.absolute())}')
        options.add_argument('--profile-directory=Default')
        
        logger.info("✓ Используется профиль для сохранения авторизации")
        logger.info(f"  Профиль: {automation_profile.absolute()}")
        logger.info("  ⚠ При первом запуске: авторизуйтесь вручную через manual_auth.py")
        logger.info("  ✓ При последующих запусках: автоматический вход")

        # Антидетект-опции для обхода защиты Wildberries
        options.add_argument("--disable-blink-features=AutomationControlled")
        options.add_argument("--start-maximized")
        options.add_argument("--disable-notifications")
        
        # Настройки скачивания (НЕ через prefs - это вызывает JSONDecodeError)
        options.add_argument(f"--download-directory={str(self.downloads_dir.absolute())}")

        # Запуск браузера (undetected-chromedriver сам скрывает WebDriver)
        logger.info("Запуск браузера...")
        
        try:
            # Запуск с правильной версией ChromeDriver для Yandex 140
            self.driver = uc.Chrome(
                options=options,
                browser_executable_path=str(browser_path),
                version_main=140,
                use_subprocess=False,
            )
            
            logger.success("✓ Браузер запущен")
            
            # КРИТИЧНО: Настройка папки скачивания через CDP (обернуто в try-except)
            try:
                self.driver.execute_cdp_cmd("Page.setDownloadBehavior", {
                    "behavior": "allow",
                    "downloadPath": str(self.downloads_dir.absolute())
                })
                logger.info(f"✓ Папка скачивания: {self.downloads_dir.absolute()}")
            except Exception as cdp_error:
                logger.warning(f"⚠ Не удалось установить папку скачивания через CDP: {cdp_error}")
                logger.info("Браузер будет использовать папку скачивания по умолчанию")
            
            # КРИТИЧНО: Даём браузеру время на полную инициализацию перед любыми действиями
            time.sleep(5)
            
        except Exception as e:
            logger.error(f"Ошибка запуска: {e}")
            raise

    def _get_yandex_browser_version(self, browser_path: Path) -> Optional[int]:
        """Определяет версию Yandex Browser.

        Args:
            browser_path: Путь к исполняемому файлу браузера

        Returns:
            Версия браузера (major version number) или None
        """
        # Если версия указана в настройках, используем её
        if self.settings.yandex_browser_version:
            logger.info(f"Используется версия из настроек: {self.settings.yandex_browser_version}")
            return self.settings.yandex_browser_version

        try:
            # Пытаемся получить версию через команду браузера
            result = subprocess.run(
                [str(browser_path), "--version"],
                capture_output=True,
                text=True,
                timeout=5,
                creationflags=subprocess.CREATE_NO_WINDOW if os.name == 'nt' else 0
            )
            
            if result.returncode == 0:
                version_output = result.stdout.strip()
                logger.debug(f"Версия браузера из --version: {version_output}")
                
                # Ищем версию в формате "Yandex Browser 138.0.7204.1908" или "138.0.7204.1908"
                match = re.search(r'(\d+)\.\d+\.\d+\.\d+', version_output)
                if match:
                    full_version = match.group(1)
                    major_version = int(full_version)
                    logger.info(f"Определена версия Yandex Browser: {major_version} (из {version_output})")
                    return major_version
        except Exception as e:
            logger.warning(f"Не удалось определить версию браузера автоматически: {e}")

        # Пытаемся найти версию в папке Application
        try:
            app_dir = browser_path.parent
            # Ищем файл с версией или папку с версией
            for item in app_dir.parent.iterdir():
                if item.is_dir() and re.match(r'^\d+\.\d+\.\d+\.\d+$', item.name):
                    major_version = int(item.name.split('.')[0])
                    logger.info(f"Определена версия Yandex Browser из папки: {major_version}")
                    return major_version
        except Exception as e:
            logger.warning(f"Не удалось определить версию из папки: {e}")

        # Если не удалось определить, используем версию по умолчанию (140)
        # Yandex Browser обычно обновляется до последних версий Chromium
        logger.warning("Не удалось определить версию браузера, используется версия по умолчанию: 140")
        return 140

    def close_browser(self) -> None:
        """Закрытие браузера."""
        if self.driver:
            try:
                self.driver.quit()
                logger.info("Браузер закрыт")
            except Exception as e:
                logger.error(f"Ошибка при закрытии браузера: {e}")

    def navigate_to_url(self, url: str) -> None:
        """Переход на указанный URL.

        Args:
            url: URL для перехода (явно указан: https://seller.wildberries.ru/analytics-reports/sales)
        """
        try:
            # Убеждаемся, что у нас есть активное окно браузера
            if not self.driver:
                raise Exception("Браузер не запущен")
            
            window_handles = self.driver.window_handles
            if not window_handles:
                raise Exception("Нет открытых окон браузера")
            
            # Переключаемся на первое окно
            self.driver.switch_to.window(window_handles[0])
            
            # Проверяем текущий URL
            try:
                current_url = self.driver.current_url
                logger.info(f"Текущий URL: {current_url}")
            except Exception:
                current_url = ""
                logger.warning("Не удалось получить текущий URL, продолжаем...")
            
            # Если уже на нужной странице, обновляем её
            if url in current_url or "seller.wildberries.ru/analytics-reports/sales" in current_url:
                logger.info("Уже на странице Wildberries, обновляем страницу")
                self.driver.refresh()
            else:
                # Открываем нужную страницу
                logger.info(f"Открытие страницы: {url}")
                self.driver.get(url)
            
            # Ждём загрузки страницы
            logger.info("Ожидание загрузки страницы...")
            time.sleep(self.settings.delay_page_load)

            # Ожидание полной загрузки страницы
            WebDriverWait(self.driver, self.settings.element_wait_timeout).until(
                lambda d: d.execute_script("return document.readyState") == "complete"
            )
            
            # Дополнительная проверка - ждём пока страница действительно загрузится
            time.sleep(2)
            
            # Проверяем текущий URL после загрузки
            try:
                final_url = self.driver.current_url
                logger.info(f"URL после загрузки: {final_url}")
                
                if url not in final_url and "seller.wildberries.ru" not in final_url:
                    logger.warning(f"⚠ Страница не открылась правильно. Текущий URL: {final_url}")
                    logger.info("Повторная попытка открытия страницы...")
                    self.driver.get(url)
                    time.sleep(self.settings.delay_page_load)
                    WebDriverWait(self.driver, self.settings.element_wait_timeout).until(
                        lambda d: d.execute_script("return document.readyState") == "complete"
                    )
                    time.sleep(2)
            except Exception as e:
                logger.warning(f"Не удалось проверить URL после загрузки: {e}")

            # Проверка авторизации
            logger.info("Проверка статуса авторизации...")
            if self._check_authorization_required():
                logger.warning("Требуется авторизация, начинаем процесс...")
                self._perform_authorization()
            else:
                logger.success("✓ Авторизация уже выполнена, пропускаем этап авторизации")

            # Проверка, что мы на странице отчётов
            # Проверяем наличие характерных элементов страницы
            page_loaded = False
            try:
                # Вариант 1: Есть поле поиска кабинетов (для пользователей с несколькими кабинетами)
                WebDriverWait(self.driver, 5).until(
                    EC.presence_of_element_located((By.ID, "suppliers-search"))
                )
                logger.success("✓ Страница отчётов загружена (найдено поле поиска кабинетов)")
                page_loaded = True
            except TimeoutException:
                # Вариант 2: Проверяем наличие кнопки календаря (есть у всех)
                try:
                    WebDriverWait(self.driver, 5).until(
                        EC.presence_of_element_located((By.CSS_SELECTOR, 'button.Date-input__icon-button__WnbzIWQzsq'))
                    )
                    logger.success("✓ Страница отчётов загружена (найдена кнопка календаря)")
                    page_loaded = True
                except TimeoutException:
                    # Вариант 3: Проверяем наличие заголовка "Продажи"
                    try:
                        WebDriverWait(self.driver, 5).until(
                            EC.presence_of_element_located((By.XPATH, "//span[text()='Продажи']"))
                        )
                        logger.success("✓ Страница отчётов загружена (найден заголовок 'Продажи')")
                        page_loaded = True
                    except TimeoutException:
                        pass
            
            if not page_loaded:
                logger.error("⚠ Не удалось найти характерные элементы страницы отчётов")
                try:
                    logger.error(f"Текущий URL: {self.driver.current_url}")
                except:
                    pass
                raise Exception("Не удалось найти элементы страницы отчётов. Возможно, требуется повторная авторизация.")

        except Exception as e:
            logger.error(f"Ошибка при переходе на страницу: {e}")
            try:
                logger.error(f"Текущий URL: {self.driver.current_url}")
            except:
                logger.error("Не удалось получить текущий URL")
            raise

    def _check_authorization_required(self) -> bool:
        """Проверка, требуется ли авторизация.

        Returns:
            True если требуется авторизация, False иначе
        """
        try:
            # Сначала проверяем, авторизованы ли мы уже (проверяем характерные элементы страницы)
            # Вариант 1: Есть поле поиска кабинетов
            try:
                WebDriverWait(self.driver, 3).until(
                    EC.presence_of_element_located((By.ID, "suppliers-search"))
                )
                logger.success("✓ Уже авторизованы - найдено поле поиска кабинетов")
                return False
            except TimeoutException:
                pass
            
            # Вариант 2: Есть кнопка календаря (характерный элемент страницы отчётов)
            try:
                WebDriverWait(self.driver, 3).until(
                    EC.presence_of_element_located((By.CSS_SELECTOR, 'button.Date-input__icon-button__WnbzIWQzsq'))
                )
                logger.success("✓ Уже авторизованы - найдена кнопка календаря на странице отчётов")
                return False
            except TimeoutException:
                pass
            
            # Вариант 3: Есть заголовок "Продажи" или "Отчеты"
            try:
                WebDriverWait(self.driver, 3).until(
                    EC.presence_of_element_located((By.XPATH, "//span[text()='Продажи' or text()='Отчеты']"))
                )
                logger.success("✓ Уже авторизованы - найден заголовок страницы отчётов")
                return False
            except TimeoutException:
                pass
            
            # Проверяем наличие поля ввода телефона (признак страницы авторизации)
            try:
                WebDriverWait(self.driver, 3).until(
                    EC.presence_of_element_located((By.CSS_SELECTOR, 'input[data-testid="phone-input"]'))
                )
                logger.warning("⚠ Требуется авторизация - обнаружено поле ввода телефона")
                return True
            except TimeoutException:
                pass
            
            # Проверяем URL - если содержит "login" или "auth", то требуется авторизация
            current_url = self.driver.current_url
            if "login" in current_url.lower() or "auth" in current_url.lower():
                logger.warning(f"⚠ Требуется авторизация - URL содержит 'login' или 'auth': {current_url}")
                return True
            
            # Если ничего не найдено, считаем что авторизованы
            logger.info("Проверка авторизации: элементы страницы авторизации не найдены, считаем что авторизованы")
            return False
            
        except Exception as e:
            logger.error(f"Ошибка при проверке авторизации: {e}")
            # В случае ошибки считаем, что авторизация требуется
            return True

    def _perform_authorization(self) -> None:
        """Выполнение авторизации."""
        try:
            if not self.settings.phone_number:
                raise Exception("Номер телефона не указан в настройках (PHONE_NUMBER)")

            logger.info("Начало процесса авторизации")

            # Шаг 1: Ввод номера телефона
            logger.info("Ввод номера телефона")
            phone_input = WebDriverWait(self.driver, self.settings.element_wait_timeout).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, 'input[data-testid="phone-input"]'))
            )
            time.sleep(self.settings.delay_before_click)  # Задержка перед кликом
            phone_input.click()
            time.sleep(self.settings.delay_after_click)  # Задержка после клика
            phone_input.clear()
            time.sleep(0.5)

            # Ввод номера телефона посимвольно
            # Убираем все символы кроме цифр
            phone_number_clean = ''.join(filter(str.isdigit, self.settings.phone_number))
            
            # Убираем +7 в начале, если есть (оставляем только цифры после +7)
            if phone_number_clean.startswith('7') and len(phone_number_clean) == 11:
                phone_number_clean = phone_number_clean[1:]  # Убираем первую 7
            elif phone_number_clean.startswith('7') and len(phone_number_clean) > 11:
                phone_number_clean = phone_number_clean[1:]  # Убираем первую 7
            
            # Вводим только цифры (без +7)
            for char in phone_number_clean:
                phone_input.send_keys(char)
                time.sleep(self.settings.delay_between_keys)

            logger.success(f"✓ Номер телефона введён: {phone_number_clean}")

            # Шаг 2: Нажатие кнопки отправки (стрелка)
            logger.info("Нажатие кнопки отправки номера")
            # Пробуем найти кнопку по разным селекторам
            try:
                submit_button = WebDriverWait(self.driver, 5).until(
                    EC.element_to_be_clickable((By.CSS_SELECTOR, 'button[data-testid="submit-phone-button"]'))
                )
            except TimeoutException:
                # Если не нашли, пробуем найти по изображению стрелки
                try:
                    submit_button = WebDriverWait(self.driver, 5).until(
                        EC.element_to_be_clickable((By.CSS_SELECTOR, 'img[alt=""][class*="FormPhoneInputBorderless__image"]'))
                    )
                except TimeoutException:
                    # Пробуем найти родительский элемент (кнопку с изображением)
                    submit_button = WebDriverWait(self.driver, 5).until(
                        EC.element_to_be_clickable((By.XPATH, '//img[contains(@class, "FormPhoneInputBorderless__image")]/parent::button | //img[contains(@class, "FormPhoneInputBorderless__image")]/ancestor::button'))
                    )
            
            time.sleep(self.settings.delay_before_click)  # Задержка перед кликом
            submit_button.click()
            time.sleep(self.settings.delay_after_click)  # Задержка после клика
            logger.success("✓ Кнопка отправки нажата")

            time.sleep(2)  # Ожидание загрузки формы ввода кода

            # Шаг 3: Запрос первого кода авторизации
            logger.info("Ожидание поля для ввода кода авторизации")
            # Ищем поле по селектору из алгоритма
            code_input = WebDriverWait(self.driver, self.settings.element_wait_timeout).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, 'div.FormTextInput__input-f6fPmoYx4c input[type="numeric"], input[type="numeric"]'))
            )

            # Запрашиваем код у пользователя
            logger.info("=" * 60)
            logger.info("ТРЕБУЕТСЯ КОД АВТОРИЗАЦИИ")
            logger.info("Пожалуйста, введите код из 6 символов, отправленный на телефон")
            logger.info("=" * 60)
            code1 = input("Введите код авторизации (6 символов): ").strip()

            # Ввод первого кода
            time.sleep(self.settings.delay_before_click)  # Задержка перед кликом
            code_input.click()
            time.sleep(self.settings.delay_after_click)  # Задержка после клика
            code_input.clear()
            time.sleep(0.5)
            for char in code1:
                code_input.send_keys(char)
                time.sleep(self.settings.delay_between_keys)

            logger.success("✓ Первый код введён")

            # Нажатие кнопки отправки (если есть)
            try:
                submit_code_button = WebDriverWait(self.driver, 3).until(
                    EC.element_to_be_clickable((By.CSS_SELECTOR, 'button[type="submit"]'))
                )
                time.sleep(self.settings.delay_before_click)  # Задержка перед кликом
                submit_code_button.click()
                time.sleep(self.settings.delay_after_click)  # Задержка после клика
                time.sleep(2)
            except TimeoutException:
                # Кнопка может отсутствовать, код может отправляться автоматически
                pass

            # Шаг 4: Запрос второго кода (код на почту)
            logger.info("Ожидание поля для ввода второго кода")
            time.sleep(2)

            # Ищем поле для второго кода (может быть то же самое или новое)
            try:
                code_input2 = WebDriverWait(self.driver, 5).until(
                    EC.presence_of_element_located((By.CSS_SELECTOR, 'div.FormTextInput__input-f6fPmoYx4c input[type="numeric"], input[type="numeric"]'))
                )
            except TimeoutException:
                # Если поле не найдено, возможно авторизация завершена
                logger.info("Поле для второго кода не найдено, проверяем авторизацию...")
                time.sleep(2)
                if not self._check_authorization_required():
                    logger.success("✓ Авторизация завершена")
                    return

            # Запрашиваем второй код у пользователя
            logger.info("=" * 60)
            logger.info("ТРЕБУЕТСЯ ВТОРОЙ КОД АВТОРИЗАЦИИ")
            logger.info("Пожалуйста, введите код из 6 символов, отправленный на почту")
            logger.info("=" * 60)
            code2 = input("Введите код авторизации с почты (6 символов): ").strip()

            # Ввод второго кода
            time.sleep(self.settings.delay_before_click)  # Задержка перед кликом
            code_input2.click()
            time.sleep(self.settings.delay_after_click)  # Задержка после клика
            code_input2.clear()
            time.sleep(0.5)
            for char in code2:
                code_input2.send_keys(char)
                time.sleep(self.settings.delay_between_keys)

            logger.success("✓ Второй код введён")

            # Нажатие кнопки входа
            try:
                login_button = WebDriverWait(self.driver, 5).until(
                    EC.element_to_be_clickable((By.CSS_SELECTOR, 'button[type="submit"]'))
                )
                time.sleep(self.settings.delay_before_click)  # Задержка перед кликом
                login_button.click()
                time.sleep(self.settings.delay_after_click)  # Задержка после клика
                logger.success("✓ Кнопка входа нажата")
            except TimeoutException:
                # Кнопка может отсутствовать, авторизация может завершиться автоматически
                pass

            # Ожидание завершения авторизации и перехода на страницу отчётов
            time.sleep(5)

            # Проверяем, что авторизация завершена
            if self._check_authorization_required():
                logger.error("Авторизация не завершена")
                raise Exception("Не удалось завершить авторизацию")

            # Переходим на страницу отчётов, если мы не на ней
            current_url = self.driver.current_url
            if "analytics-reports/sales" not in current_url:
                logger.info("Переход на страницу отчётов после авторизации")
                self.driver.get(self.settings.wildberries_start_url)
                time.sleep(self.settings.delay_page_load)

            logger.success("✓ Авторизация завершена")

        except Exception as e:
            logger.error(f"Ошибка при авторизации: {e}")
            logger.exception("Детали ошибки:")
            raise

    def wait_for_element(self, by: By, value: str, timeout: Optional[int] = None) -> None:
        """Ожидание появления элемента на странице.

        Args:
            by: Тип селектора (By.ID, By.CLASS_NAME и т.д.)
            value: Значение селектора
            timeout: Таймаут ожидания (по умолчанию из настроек)
        """
        timeout = timeout or self.settings.element_wait_timeout
        try:
            WebDriverWait(self.driver, timeout).until(
                EC.presence_of_element_located((by, value))
            )
        except TimeoutException:
            logger.error(f"Элемент не найден: {by}={value}")
            raise

    def click_element(self, by: By, value: str, scroll: bool = True) -> None:
        """Клик по элементу.

        Args:
            by: Тип селектора
            value: Значение селектора
            scroll: Прокрутить страницу к элементу перед кликом
        """
        try:
            time.sleep(self.settings.delay_before_click)
            element = WebDriverWait(self.driver, self.settings.element_wait_timeout).until(
                EC.element_to_be_clickable((by, value))
            )

            if scroll:
                self.driver.execute_script("arguments[0].scrollIntoView(true);", element)
                time.sleep(0.5)

            element.click()
            time.sleep(self.settings.delay_after_click)
            logger.debug(f"Клик по элементу: {by}={value}")

        except Exception as e:
            logger.error(f"Ошибка при клике по элементу {by}={value}: {e}")
            raise

    def fill_input(self, by: By, value: str, text: str, clear: bool = True, scroll: bool = True) -> None:
        """Заполнение поля ввода.

        Args:
            by: Тип селектора
            value: Значение селектора
            text: Текст для ввода
            clear: Очистить поле перед вводом
            scroll: Прокрутить страницу к элементу перед вводом
        """
        # Повторные попытки при StaleElementReferenceException
        max_retries = 3
        for attempt in range(max_retries):
            try:
                time.sleep(self.settings.delay_before_type)
                
                # ВСЕГДА ищем элемент заново (защита от stale element)
                element = WebDriverWait(self.driver, self.settings.element_wait_timeout).until(
                    EC.presence_of_element_located((by, value))
                )

                if scroll:
                    self.driver.execute_script("arguments[0].scrollIntoView(true);", element)
                    time.sleep(0.5)

                if clear:
                    # Используем JavaScript для очистки (надёжнее)
                    self.driver.execute_script("arguments[0].value = '';", element)
                    time.sleep(0.3)

                # Посимвольный ввод для имитации человеческого поведения
                for char in text:
                    element.send_keys(char)
                    time.sleep(self.settings.delay_between_keys)

                time.sleep(self.settings.delay_after_type)
                logger.debug(f"Заполнено поле {by}={value}: {text}")
                return  # Успешно - выходим

            except StaleElementReferenceException as e:
                if attempt < max_retries - 1:
                    logger.warning(f"⚠ Элемент устарел, повтор {attempt + 1}/{max_retries}...")
                    time.sleep(1)
                    continue
                else:
                    logger.error(f"Ошибка при заполнении поля {by}={value}: {e}")
                    raise
            except Exception as e:
                logger.error(f"Ошибка при заполнении поля {by}={value}: {e}")
                raise

    def delete_all_reports(self) -> None:
        """Удаляет все отчёты на странице перед скачиванием.
        
        Находит все кнопки удаления (с иконкой корзины) и нажимает на них.
        """
        try:
            logger.info("   Поиск кнопок удаления отчётов...")
            
            # Ждём загрузки страницы
            time.sleep(2)
            
            # Ищем все кнопки удаления по селектору из HTML
            # Кнопка содержит SVG с path для корзины (более стабильный способ поиска)
            delete_buttons = []
            
            # Вариант 1: Ищем по SVG path напрямую (самый надёжный способ)
            try:
                # Ищем все SVG с path для корзины (path содержит характерную последовательность для иконки корзины)
                # Из HTML: path с d="M7 0H13C14.1046 0 15 0.89543 15 2V3H18C19.1046 3 20 3.89543 20 5V7..."
                svg_paths = self.driver.find_elements(
                    By.XPATH,
                    '//svg//path[contains(@d, "M7 0H13") or contains(@d, "M17 0H13")]'
                )
                for svg_path in svg_paths:
                    try:
                        # Находим родительскую кнопку
                        button = svg_path.find_element(By.XPATH, './ancestor::button[1]')
                        if button not in delete_buttons:
                            delete_buttons.append(button)
                    except:
                        continue
                if delete_buttons:
                    logger.debug(f"Найдено кнопок удаления через SVG path: {len(delete_buttons)}")
            except Exception as e:
                logger.warning(f"Не удалось найти кнопки удаления по SVG path: {e}")
            
            # Вариант 2: Ищем по классу кнопки (резервный способ)
            if not delete_buttons:
                try:
                    buttons = self.driver.find_elements(
                        By.CSS_SELECTOR,
                        'button[type="button"][class*="Button-link"]'
                    )
                    # Проверяем, что кнопка содержит иконку корзины
                    for button in buttons:
                        try:
                            svg = button.find_element(By.TAG_NAME, "svg")
                            path = svg.find_element(By.TAG_NAME, "path")
                            path_d = path.get_attribute("d")
                            # Проверяем наличие характерного path для корзины
                            if path_d and ("M7 0H13" in path_d or "M17 0H13" in path_d or "d=\"M7 0H13" in path_d):
                                if button not in delete_buttons:
                                    delete_buttons.append(button)
                        except:
                            continue
                    if delete_buttons:
                        logger.debug(f"Найдено кнопок удаления через класс: {len(delete_buttons)}")
                except Exception as e:
                    logger.warning(f"Не удалось найти кнопки удаления по классу: {e}")
            
            # Вариант 3: Ищем все кнопки с иконкой корзины по aria-label или title
            if not delete_buttons:
                try:
                    # Ищем кнопки, которые могут иметь подсказку об удалении
                    buttons = self.driver.find_elements(
                        By.XPATH,
                        '//button[contains(@aria-label, "удал") or contains(@title, "удал") or contains(@aria-label, "delete") or contains(@title, "delete")]'
                    )
                    for button in buttons:
                        try:
                            # Проверяем, что это действительно кнопка удаления (содержит SVG корзины)
                            svg = button.find_element(By.TAG_NAME, "svg")
                            if button not in delete_buttons:
                                delete_buttons.append(button)
                        except:
                            continue
                    if delete_buttons:
                        logger.debug(f"Найдено кнопок удаления через aria-label: {len(delete_buttons)}")
                except Exception as e:
                    logger.warning(f"Не удалось найти кнопки удаления по aria-label: {e}")
            
            if not delete_buttons:
                logger.info("   ✓ Кнопки удаления не найдены (отчётов нет на странице)")
                return
            
            logger.info(f"   Найдено кнопок удаления: {len(delete_buttons)}")
            logger.info("   Начинаем удаление отчётов...")
            
            # Нажимаем на каждую кнопку удаления
            deleted_count = 0
            for i, button in enumerate(delete_buttons, 1):
                try:
                    logger.info(f"     Удаляем отчёт {i}/{len(delete_buttons)}...")
                    # Прокручиваем к кнопке
                    self.driver.execute_script("arguments[0].scrollIntoView(true);", button)
                    time.sleep(0.5)
                    
                    # Нажимаем на кнопку
                    time.sleep(self.settings.delay_before_click)
                    button.click()
                    time.sleep(self.settings.delay_after_click)
                    deleted_count += 1
                    logger.info(f"     ✓ Отчёт {i} удалён")
                    
                    # Небольшая задержка между удалениями
                    time.sleep(1)
                except Exception as e:
                    logger.warning(f"     ⚠ Не удалось удалить отчёт {i}: {e}")
                    continue
            
            logger.success(f"   ✅ Удалено отчётов: {deleted_count}/{len(delete_buttons)}")
            
            # Ждём обновления страницы после удаления
            logger.info("   Ожидание обновления страницы...")
            time.sleep(2)
            
        except Exception as e:
            logger.error(f"Ошибка при удалении отчётов: {e}")
            logger.exception("Детали ошибки:")
            # Не прерываем выполнение, продолжаем работу

    def process_cabinet(self, cabinet: Dict[str, str], target_date: Optional[date] = None) -> Optional[Path]:
        """Обработка одного кабинета.

        Args:
            cabinet: Словарь с информацией о кабинете (name, id)
            target_date: Дата для скачивания отчёта (если None, используется вчерашний день)

        Returns:
            Путь к обработанному файлу или None в случае ошибки
        """
        cabinet_name = cabinet["name"]
        cabinet_id = cabinet["id"]
        
        # Определяем дату для скачивания
        if target_date:
            report_date = target_date
        else:
            report_date = (datetime.now() - timedelta(days=1)).date()
        
        date_str = report_date.strftime("%d.%m.%Y")

        logger.info("=" * 70)
        logger.info(f"📋 НАЧАЛО ОБРАБОТКИ КАБИНЕТА: {cabinet_name.upper()}")
        logger.info(f"   ID кабинета: {cabinet_id}")
        logger.info(f"   Дата отчёта: {date_str}")
        logger.info("=" * 70)

        try:
            # Шаг 2.1: Раскрытие меню выбора кабинетов
            logger.info("")
            logger.info("🔹 ШАГ 1: Раскрытие меню выбора кабинетов")
            logger.info("   Ищем кнопку выбора кабинетов...")
            try:
                # Ищем кнопку с именем пользователя/кабинета (содержит стрелку вниз)
                # Используем data-testid для надёжности
                dropdown_button = WebDriverWait(self.driver, 5).until(
                    EC.element_to_be_clickable((By.CSS_SELECTOR, 'button[data-testid="desktop-profile-select-button-chips-component"]'))
                )
                logger.info("   ✓ Кнопка найдена, кликаем...")
                time.sleep(self.settings.delay_before_click)
                dropdown_button.click()
                time.sleep(self.settings.delay_after_click)
                logger.success("   ✅ Меню выбора кабинетов раскрыто")
            except TimeoutException:
                logger.warning("   ⚠ Кнопка раскрытия меню не найдена, возможно меню уже раскрыто")
            
            # Шаг 2.2: Ввод ID кабинета
            logger.info("")
            logger.info("🔹 ШАГ 2: Поиск и выбор кабинета")
            logger.info(f"   Ищем кабинет с ID: {cabinet_id}")
            
            # КРИТИЧНО: Ждём появления поля поиска после раскрытия меню
            try:
                logger.info("   Ожидание появления поля поиска...")
                WebDriverWait(self.driver, 10).until(
                    EC.presence_of_element_located((By.ID, "suppliers-search"))
                )
                logger.info("   ✓ Поле поиска появилось")
                time.sleep(1)  # Дополнительная задержка для стабильности
            except TimeoutException:
                logger.warning("   ⚠ Поле поиска не появилось, пробуем продолжить...")
            
            logger.info(f"   Вводим ID кабинета: {cabinet_id}")
            self.fill_input(
                By.ID,
                "suppliers-search",
                cabinet_id,
                clear=True
            )
            logger.info("   ✓ ID введён, ждём результатов поиска...")
            time.sleep(2)  # Ожидание загрузки результатов поиска
            
            # КРИТИЧНО: Нажимаем на найденный кабинет
            logger.info(f"   Кликаем на найденный кабинет {cabinet_id}...")
            try:
                # Вариант 1: Пробуем кликнуть на label кабинета
                try:
                    cabinet_label = WebDriverWait(self.driver, 5).until(
                        EC.element_to_be_clickable((By.CSS_SELECTOR, 'label.suppliers-item-new_SuppliersItem__label__j6lv6'))
                    )
                    time.sleep(self.settings.delay_before_click)
                    cabinet_label.click()
                    time.sleep(self.settings.delay_after_click)
                    logger.success(f"   ✅ Кабинет {cabinet_id} выбран (через label)")
                except:
                    # Вариант 2: Пробуем кликнуть на checkbox label
                    try:
                        checkbox_label = WebDriverWait(self.driver, 5).until(
                            EC.element_to_be_clickable((By.CSS_SELECTOR, 'label[data-testid="supplier-checkbox-checkbox"]'))
                        )
                        time.sleep(self.settings.delay_before_click)
                        checkbox_label.click()
                        time.sleep(self.settings.delay_after_click)
                        logger.success(f"   ✅ Кабинет {cabinet_id} выбран (через checkbox label)")
                    except:
                        # Вариант 3: Последняя попытка - кликаем на сам input
                        cabinet_input = WebDriverWait(self.driver, 5).until(
                            EC.element_to_be_clickable((By.CSS_SELECTOR, 'input[data-testid="supplier-checkbox-checkbox-input"]'))
                        )
                        time.sleep(self.settings.delay_before_click)
                        cabinet_input.click()
                        time.sleep(self.settings.delay_after_click)
                        logger.success(f"   ✅ Кабинет {cabinet_id} выбран (через input)")
            except Exception as e:
                logger.warning(f"   ⚠ Не удалось кликнуть на кабинет {cabinet_id}: {e}, продолжаем...")

            # Шаг 2.2.5: Удаление всех отчётов перед скачиванием
            logger.info("")
            logger.info("🔹 ШАГ 3: Удаление всех существующих отчётов")
            logger.info("   Ищем и удаляем все отчёты в кабинете...")
            self.delete_all_reports()
            
            # Шаг 2.3: Настройка периода отчёта
            logger.info("")
            logger.info("🔹 ШАГ 4: Настройка периода отчёта")
            logger.info(f"   Устанавливаем дату: {date_str}")
            logger.info("   Ищем кнопку календаря...")
            self.click_element(By.CSS_SELECTOR, "button.Date-input__icon-button__WnbzIWQzsq")
            logger.info("   ✓ Кнопка календаря нажата")

            # Ожидание появления календаря и полей ввода даты
            logger.info("   Ожидание появления полей ввода даты...")
            try:
                WebDriverWait(self.driver, self.settings.element_wait_timeout).until(
                    EC.presence_of_element_located((By.ID, "startDate"))
                )
                WebDriverWait(self.driver, self.settings.element_wait_timeout).until(
                    EC.presence_of_element_located((By.ID, "endDate"))
                )
                logger.info("   ✓ Поля ввода даты найдены")
            except TimeoutException:
                logger.error("   ❌ Поля ввода даты не найдены")
                raise

            time.sleep(0.5)

            # Заполнение поля начала периода
            logger.info(f"   Заполняем поле 'Начало периода': {date_str}")
            start_date_input = WebDriverWait(self.driver, self.settings.element_wait_timeout).until(
                EC.element_to_be_clickable((By.ID, "startDate"))
            )
            time.sleep(self.settings.delay_before_click)  # Задержка перед кликом
            start_date_input.click()
            time.sleep(self.settings.delay_after_click)  # Задержка после клика
            start_date_input.clear()
            time.sleep(0.5)
            for char in date_str:
                start_date_input.send_keys(char)
                time.sleep(self.settings.delay_between_keys)
            logger.info("   ✓ Поле 'Начало периода' заполнено")

            # Заполнение поля окончания периода
            logger.info(f"   Заполняем поле 'Конец периода': {date_str}")
            end_date_input = WebDriverWait(self.driver, self.settings.element_wait_timeout).until(
                EC.element_to_be_clickable((By.ID, "endDate"))
            )
            time.sleep(self.settings.delay_before_click)  # Задержка перед кликом
            end_date_input.click()
            time.sleep(self.settings.delay_after_click)  # Задержка после клика
            end_date_input.clear()
            time.sleep(0.5)
            for char in date_str:
                end_date_input.send_keys(char)
                time.sleep(self.settings.delay_between_keys)
            logger.info("   ✓ Поле 'Конец периода' заполнено")

            # Нажатие кнопки "Сохранить"
            logger.info("   Ищем кнопку 'Сохранить'...")
            # Используем более точный селектор с текстом "Сохранить"
            save_button = WebDriverWait(self.driver, self.settings.element_wait_timeout).until(
                EC.element_to_be_clickable((By.XPATH, "//button[@type='submit' and .//span[text()='Сохранить']]"))
            )
            time.sleep(self.settings.delay_before_click)  # Задержка перед кликом
            save_button.click()
            time.sleep(self.settings.delay_after_click)  # Задержка после клика
            logger.success("   ✅ Период сохранён")

            # Шаг 2.4: Выгрузка отчёта
            logger.info("")
            logger.info("🔹 ШАГ 5: Выгрузка отчёта в Excel")
            logger.info("   Ожидание после сохранения периода...")
            time.sleep(3)  # Ожидание после сохранения периода

            # Очищаем папку downloads перед скачиванием (чтобы найти только новый файл)
            logger.info("   Очищаем папку downloads от старых файлов...")
            self._clear_downloads_folder()
            
            # Поиск кнопки "Выгрузить в Excel"
            logger.info("   Ищем кнопку 'Выгрузить в Excel'...")
            download_button = WebDriverWait(self.driver, self.settings.element_wait_timeout).until(
                EC.element_to_be_clickable((By.XPATH, "//button[.//span[text()='Выгрузить в Excel']]"))
            )
            logger.info("   ✓ Кнопка найдена, нажимаем...")
            time.sleep(self.settings.delay_before_click)  # Задержка перед кликом
            download_button.click()
            time.sleep(self.settings.delay_after_click)  # Задержка после клика
            logger.success("   ✅ Запрос на выгрузку отправлен")

            # Ожидание скачивания файла
            logger.info("   Ожидание скачивания файла...")
            downloaded_file = self._wait_for_downloaded_file()
            if not downloaded_file:
                logger.error("   ❌ Файл не был скачан")
                return None
            logger.info(f"   ✓ Файл скачан: {downloaded_file.name}")

            # Шаг 2.4: Обработка скачанного файла
            logger.info("")
            logger.info("🔹 ШАГ 6: Обработка скачанного файла")
            logger.info(f"   Переименовываем файл в: {cabinet_name} {date_str}.xlsx")
            processed_file = self._process_downloaded_file(downloaded_file, cabinet_name, date_str)
            if not processed_file:
                logger.error("   ❌ Ошибка при обработке файла")
                return None
            logger.success(f"   ✅ Файл обработан: {processed_file.name}")

            # Шаг 2.5: Создание резервной копии
            logger.info("")
            logger.info("🔹 ШАГ 7: Создание резервной копии")
            logger.info(f"   Сохраняем копию в папку data/{date_str}/...")
            backup_file = self._create_backup(processed_file, cabinet_name, date_str)
            if not backup_file:
                logger.error("   ❌ Ошибка при создании резервной копии")
                return None
            logger.success(f"   ✅ Резервная копия создана: {backup_file.name}")

            logger.info("")
            logger.success("=" * 70)
            logger.success(f"✅ КАБИНЕТ {cabinet_name.upper()} УСПЕШНО ОБРАБОТАН")
            logger.success(f"   Файл: {processed_file.name}")
            logger.success("=" * 70)
            return processed_file

        except Exception as e:
            logger.error(f"Ошибка при обработке кабинета {cabinet_name}: {e}")
            logger.exception("Детали ошибки:")
            return None

    def _clear_downloads_folder(self) -> None:
        """Очищает папку downloads от старых файлов перед скачиванием."""
        try:
            files_before = list(self.downloads_dir.glob("*.xlsx")) + list(self.downloads_dir.glob("*.xls"))
            if files_before:
                logger.info(f"   Найдено старых файлов: {len(files_before)}")
                for old_file in files_before:
                    try:
                        old_file.unlink()
                        logger.debug(f"   Удалён старый файл: {old_file.name}")
                    except Exception as e:
                        logger.warning(f"   Не удалось удалить {old_file.name}: {e}")
                logger.info("   ✓ Папка downloads очищена")
            else:
                logger.info("   ✓ Папка downloads уже пуста")
        except Exception as e:
            logger.warning(f"   ⚠ Ошибка при очистке папки downloads: {e}")

    def _wait_for_downloaded_file(self, timeout: int = 60) -> Optional[Path]:
        """Ожидание скачивания файла.

        Args:
            timeout: Таймаут ожидания в секундах

        Returns:
            Путь к скачанному файлу или None
        """
        start_time = time.time()
        logger.info(f"   Ожидание до {timeout} секунд...")
        check_count = 0
        
        while time.time() - start_time < timeout:
            check_count += 1
            if check_count % 5 == 0:  # Каждые 5 секунд показываем прогресс
                elapsed = int(time.time() - start_time)
                logger.info(f"   Ожидание... ({elapsed}/{timeout} сек)")
            
            # Ищем файлы .xlsx и .xls в папке downloads
            for file_path in self.downloads_dir.glob("*.xlsx"):
                # Проверяем, что файл не заблокирован (завершено скачивание)
                try:
                    if file_path.stat().st_size > 0:
                        # Проверяем, что файл не был изменён недавно (скачивание завершено)
                        if time.time() - file_path.stat().st_mtime > 2:
                            logger.info(f"   ✓ Найден новый файл: {file_path.name} ({file_path.stat().st_size} байт)")
                            return file_path
                except Exception:
                    pass

            for file_path in self.downloads_dir.glob("*.xls"):
                try:
                    if file_path.stat().st_size > 0:
                        if time.time() - file_path.stat().st_mtime > 2:
                            logger.info(f"   ✓ Найден новый файл: {file_path.name} ({file_path.stat().st_size} байт)")
                            return file_path
                except Exception:
                    pass

            time.sleep(1)

        logger.error(f"   ❌ Таймаут ожидания скачивания файла ({timeout} сек)")
        return None

    def _process_downloaded_file(
        self, file_path: Path, cabinet_name: str, date_str: str
    ) -> Optional[Path]:
        """Обработка скачанного файла: переименование и замена первой строки.

        Args:
            file_path: Путь к скачанному файлу
            cabinet_name: Название кабинета (используется для имени файла)
            date_str: Дата в формате DD.MM.YYYY

        Returns:
            Путь к обработанному файлу или None
        """
        try:
            # Новое имя файла - используем имя кабинета как есть
            new_name = f"{cabinet_name} {date_str}.xlsx"
            new_path = self.downloads_dir / new_name

            logger.info(f"   Исходный файл: {file_path.name}")
            logger.info(f"   Новое имя: {new_name}")

            # Если файл с таким именем уже существует, удаляем его
            if new_path.exists() and file_path != new_path:
                logger.warning(f"   ⚠ Файл {new_name} уже существует, будет перезаписан")
                new_path.unlink()

            # Переименование файла
            if file_path != new_path:
                file_path.rename(new_path)
                logger.info(f"   ✓ Файл переименован: {file_path.name} → {new_name}")

            # Замена первой строки
            logger.info("   Обрабатываем заголовки файла...")
            logger.info("     - Разъединение объединённых ячеек...")
            logger.info("     - Удаление первой строки...")
            logger.info("     - Вставка правильных заголовков...")
            self._replace_first_row(new_path)
            logger.info("   ✓ Заголовки обработаны")

            return new_path

        except Exception as e:
            logger.error(f"   ❌ Ошибка при обработке файла: {e}")
            logger.exception("Детали ошибки:")
            return None

    def _replace_first_row(self, file_path: Path) -> None:
        """Удаление первой строки и замена заголовков на жёстко заданные значения.

        Алгоритм:
        1. Разъединяем объединённые ячейки в первой строке
        2. Удаляем первую строку (неполные заголовки)
        3. Заменяем новую первую строку (бывшую вторую) на жёстко заданные заголовки

        Args:
            file_path: Путь к файлу
        """
        try:
            # Жёстко заданные заголовки (A-P столбцы)
            headers = [
                "Бренд",           # A
                "Предмет",         # B
                "Сезон",           # C
                "Коллекция",       # D
                "Наименование",    # E
                "Артикул поставщика",  # F
                "Номенклатура",    # G
                "Баркод",          # H
                "Размер",          # I
                "Контракт",        # J
                "Склад",           # K
                "Заказано шт",     # L
                "Заказано себестоимость",  # M
                "Выкупили шт",     # N
                "Выкупили руб",    # O
                "Текущий остаток"  # P
            ]

            # Загружаем файл для обработки
            wb = load_workbook(file_path)
            ws = wb.active

            # КРИТИЧНО: Шаг 1: Разъединяем все объединённые ячейки в первой строке
            # (WB выгружает файлы с объединёнными ячейками в заголовках)
            logger.debug("Разъединение объединённых ячеек в первой строке...")
            merged_ranges = list(ws.merged_cells.ranges)
            for merged_range in merged_ranges:
                # Проверяем, относится ли объединение к первой строке
                if merged_range.min_row == 1 and merged_range.max_row == 1:
                    logger.debug(f"  Разъединяем: {merged_range}")
                    ws.unmerge_cells(str(merged_range))

            # Шаг 2: Удаляем первую строку (неполные заголовки)
            logger.debug("Удаление первой строки...")
            ws.delete_rows(1)

            # Шаг 3: Заменяем новую первую строку (бывшую вторую) на жёстко заданные заголовки
            logger.debug("Замена заголовков на жёстко заданные значения...")
            for col_idx, header in enumerate(headers, start=1):
                ws.cell(row=1, column=col_idx).value = header

            # Сохраняем изменения
            wb.save(file_path)
            wb.close()
            logger.success("✓ Первая строка удалена, заголовки заменены на жёстко заданные значения")

        except Exception as e:
            logger.error(f"Ошибка при замене первой строки: {e}")
            logger.exception("Детали ошибки:")
            raise

    def _create_backup(self, file_path: Path, cabinet_name: str, date_str: str) -> Optional[Path]:
        """Создание резервной копии файла в папке data.

        Args:
            file_path: Путь к исходному файлу
            cabinet_name: Название кабинета
            date_str: Дата в формате DD.MM.YYYY

        Returns:
            Путь к резервной копии или None
        """
        try:
            # Создаём папку с датой
            date_folder = self.data_dir / date_str
            date_folder.mkdir(parents=True, exist_ok=True)

            # Новое имя файла для резервной копии
            backup_name = f"{cabinet_name.lower()}_{date_str}.xlsx"
            backup_path = date_folder / backup_name

            # Копируем файл
            shutil.copy2(file_path, backup_path)
            logger.success(f"✓ Резервная копия создана: {backup_path}")

            return backup_path

        except Exception as e:
            logger.error(f"Ошибка при создании резервной копии: {e}")
            logger.exception("Детали ошибки:")
            return None

    def _detect_current_page_state(self) -> str:
        """Определяет на какой странице мы сейчас находимся.
        
        Returns:
            'auth_required' - страница авторизации
            'reports_page' - страница отчётов (готова к работе)
            'unknown' - неизвестная страница
        """
        try:
            current_url = self.driver.current_url
            logger.debug(f"Текущий URL: {current_url}")
            
            # Проверка 1: Страница авторизации
            if "seller-auth.wildberries.ru" in current_url:
                logger.debug("→ Обнаружена страница авторизации")
                return "auth_required"
            
            # Проверка 2: Страница отчётов - ищем характерные элементы
            try:
                # Ищем любой из характерных элементов страницы отчётов
                self.driver.find_element(By.ID, "suppliers-search")
                logger.debug("→ Обнаружена страница отчётов (поле поиска)")
                return "reports_page"
            except:
                pass
            
            try:
                self.driver.find_element(By.CSS_SELECTOR, 'button.Date-input__icon-button__WnbzIWQzsq')
                logger.debug("→ Обнаружена страница отчётов (кнопка календаря)")
                return "reports_page"
            except:
                pass
            
            try:
                self.driver.find_element(By.XPATH, "//span[text()='Продажи']")
                logger.debug("→ Обнаружена страница отчётов (заголовок)")
                return "reports_page"
            except:
                pass
            
            # Если ничего не нашли
            logger.debug("→ Страница не распознана")
            return "unknown"
            
        except Exception as e:
            logger.error(f"Ошибка при определении состояния страницы: {e}")
            return "unknown"

    def execute_flow(self, target_date: Optional[date] = None) -> None:
        """Выполнение основного потока работы для всех кабинетов.
        
        Args:
            target_date: Дата для скачивания отчётов (если None, используется вчерашний день)
        """
        try:
            # Запуск браузера
            self.start_browser()
            
            # Ждём стабилизации браузера
            logger.info("Ожидание стабилизации браузера...")
            time.sleep(3)
            
            # Открываем страницу Wildberries
            logger.info(f"Открытие страницы {self.WILDBERRIES_REPORTS_URL}...")
            self.driver.get(self.WILDBERRIES_REPORTS_URL)
            
            # Ждём загрузки страницы
            logger.info("Ожидание загрузки страницы...")
            time.sleep(self.settings.delay_page_load)
            
            # === УМНЫЙ ЦИКЛ ПРОВЕРКИ СОСТОЯНИЯ ===
            # Проверяем состояние страницы каждые 30 секунд и выполняем нужные действия
            logger.info("=" * 60)
            logger.info("ЗАПУСК УМНОГО МОНИТОРИНГА СОСТОЯНИЯ")
            logger.info("Скрипт будет проверять состояние страницы каждые 30 секунд")
            logger.info("=" * 60)
            
            max_wait_cycles = 10  # Максимум 10 циклов по 30 секунд = 5 минут ожидания
            current_cycle = 0
            authorized = False
            
            while not authorized and current_cycle < max_wait_cycles:
                current_cycle += 1
                logger.info(f"[Цикл {current_cycle}/{max_wait_cycles}] Проверка состояния страницы...")
                
                # Определяем состояние
                page_state = self._detect_current_page_state()
                
                if page_state == "auth_required":
                    logger.info("=" * 60)
                    logger.info("ОБНАРУЖЕНА СТРАНИЦА АВТОРИЗАЦИИ")
                    logger.info("Запуск процесса авторизации...")
                    logger.info("=" * 60)
                    try:
                        self._perform_authorization()
                        # После авторизации ждём загрузки следующей страницы
                        time.sleep(5)
                        # Проверяем, попали ли на страницу отчётов
                        if self._detect_current_page_state() == "reports_page":
                            logger.success("✓ АВТОРИЗАЦИЯ УСПЕШНА! Переход на страницу отчётов")
                            authorized = True
                            break
                        else:
                            logger.warning("⚠ После авторизации не попали на страницу отчётов, продолжаем ждать...")
                    except Exception as e:
                        logger.error(f"Ошибка при авторизации: {e}")
                        logger.info("Продолжаем мониторинг...")
                    
                elif page_state == "reports_page":
                    logger.success("=" * 60)
                    logger.success("✓ ОБНАРУЖЕНА СТРАНИЦА ОТЧЁТОВ")
                    logger.success("Авторизация не требуется, начинаем работу с кабинетами")
                    logger.success("=" * 60)
                    authorized = True
                    break
                    
                else:
                    logger.warning(f"⚠ Страница не распознана, ожидание 30 секунд...")
                    logger.info(f"Текущий URL: {self.driver.current_url}")
                    time.sleep(30)
            
            if not authorized:
                logger.error("✗ НЕ УДАЛОСЬ АВТОРИЗОВАТЬСЯ ИЛИ ПОПАСТЬ НА СТРАНИЦУ ОТЧЁТОВ")
                logger.error(f"Истекло время ожидания ({max_wait_cycles * 30} секунд)")
                raise Exception("Не удалось авторизоваться")
            
            # === РАБОТА С КАБИНЕТАМИ (запускается только если authorized=True) ===
            
            # Раскрытие меню выбора кабинетов на главной странице
            logger.info("Раскрытие меню выбора кабинетов на главной странице...")
            try:
                # Ищем кнопку с именем пользователя/кабинета для раскрытия меню
                profile_button = WebDriverWait(self.driver, 10).until(
                    EC.element_to_be_clickable((By.CSS_SELECTOR, 'button[data-testid="desktop-profile-select-button-chips-component"]'))
                )
                time.sleep(self.settings.delay_before_click)
                profile_button.click()
                time.sleep(self.settings.delay_after_click)
                logger.success("✓ Меню выбора кабинетов раскрыто на главной странице")
            except TimeoutException:
                logger.warning("⚠ Кнопка раскрытия меню не найдена, возможно меню уже раскрыто или у пользователя один кабинет")
            except Exception as e:
                logger.warning(f"⚠ Ошибка при раскрытии меню: {e}, продолжаем работу...")

            # Обработка каждого кабинета
            total_cabinets = len(self.CABINETS)
            logger.info("")
            logger.info("=" * 70)
            logger.info(f"📊 НАЧИНАЕМ ОБРАБОТКУ {total_cabinets} КАБИНЕТОВ")
            logger.info("=" * 70)
            
            for idx, cabinet in enumerate(self.CABINETS, 1):
                try:
                    logger.info("")
                    logger.info("")
                    logger.info("╔" + "═" * 68 + "╗")
                    logger.info(f"║  КАБИНЕТ {idx}/{total_cabinets}: {cabinet['name'].upper()} (ID: {cabinet['id']})")
                    logger.info("╚" + "═" * 68 + "╝")
                    
                    # Проверка состояния перед обработкой кабинета
                    logger.info("Проверка состояния страницы...")
                    page_state = self._detect_current_page_state()
                    
                    if page_state == "auth_required":
                        logger.warning("⚠ Требуется повторная авторизация!")
                        self._perform_authorization()
                        time.sleep(5)
                        # Переход обратно на страницу отчётов
                        logger.info("Переход на страницу отчётов...")
                        self.driver.get(self.settings.wildberries_start_url)
                        time.sleep(self.settings.delay_page_load)
                    elif page_state == "unknown":
                        logger.warning("⚠ Неизвестная страница, переход на страницу отчётов...")
                        self.driver.get(self.settings.wildberries_start_url)
                        time.sleep(self.settings.delay_page_load)
                    else:
                        logger.info("✓ Страница отчётов доступна")
                    
                    # Обработка кабинета
                    result = self.process_cabinet(cabinet, target_date=target_date)

                    if result:
                        logger.success(f"✅ Кабинет {cabinet['name']} обработан успешно")
                        logger.info(f"   Файл сохранён: {result.name}")
                    else:
                        logger.error(f"❌ Ошибка при обработке кабинета {cabinet['name']}")

                    # Возврат на стартовую страницу для следующего кабинета
                    if cabinet != self.CABINETS[-1]:  # Не возвращаемся после последнего кабинета
                        logger.info("")
                        logger.info("⏭ Переход к следующему кабинету...")
                        logger.info("   Возврат на стартовую страницу...")
                        self.driver.get(self.settings.wildberries_start_url)
                        
                        # Ждём полной загрузки страницы
                        logger.info("   Ожидание загрузки страницы...")
                        time.sleep(self.settings.delay_page_load)
                        
                        # КРИТИЧНО: Заново раскрываем меню для следующего кабинета
                        logger.info("   Раскрытие меню для следующего кабинета...")
                        try:
                            profile_button = WebDriverWait(self.driver, 10).until(
                                EC.element_to_be_clickable((By.CSS_SELECTOR, 'button[data-testid="desktop-profile-select-button-chips-component"]'))
                            )
                            time.sleep(self.settings.delay_before_click)
                            profile_button.click()
                            time.sleep(self.settings.delay_after_click)
                            logger.info("   ✓ Меню раскрыто")
                        except TimeoutException:
                            logger.warning("   ⚠ Кнопка раскрытия меню не найдена")
                        except Exception as e:
                            logger.warning(f"   ⚠ Ошибка при раскрытии меню: {e}")

                except Exception as e:
                    logger.error(f"❌ Критическая ошибка при обработке кабинета {cabinet['name']}: {e}")
                    logger.exception("Детали ошибки:")
                    continue

            logger.info("")
            logger.info("=" * 70)
            logger.success("✅ ВСЕ КАБИНЕТЫ ОБРАБОТАНЫ")
            logger.info("=" * 70)

        except Exception as e:
            logger.error(f"Критическая ошибка в процессе выполнения: {e}")
            logger.exception("Детали ошибки:")
            raise

        finally:
            # Закрытие браузера
            self.close_browser()
