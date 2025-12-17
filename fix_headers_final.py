"""Финальная версия скрипта для исправления заголовков."""
from pathlib import Path
from openpyxl import load_workbook

# Правильные заголовки по столбцам A-P
CORRECT_HEADERS = [
    "Бренд",           # A
    "Предмет",         # B
    "Сезон",           # C
    "Коллекция",       # D
    "Наименование",    # E
    "Артикул поставщика",  # F
    "Номенклатура",    # G
    "Баркод",          # H
    "Размер",          # I
    "Контракт",        # J
    "Склад",           # K
    "Заказано шт",     # L
    "Заказано себестоимость",  # M
    "Выкупили шт",     # N
    "Выкупили руб",    # O
    "Текущий остаток"  # P
]

def fix_file_headers(file_path: Path):
    """Исправляет заголовки в файле.
    
    1. Разъединяет объединённые ячейки в первой строке
    2. Удаляет первую строку
    3. Записывает правильные заголовки в новую первую строку
    """
    print(f"\n{'='*60}")
    print(f"Обработка: {file_path.name}")
    print('='*60)
    
    # Загружаем файл
    wb = load_workbook(file_path)
    ws = wb.active
    
    # Шаг 1: Разъединяем все объединённые ячейки в первой строке
    print("\n>>> Разъединение объединённых ячеек...")
    merged_ranges = list(ws.merged_cells.ranges)
    for merged_range in merged_ranges:
        # Проверяем, относится ли объединение к первой строке
        if merged_range.min_row == 1 and merged_range.max_row == 1:
            print(f"  Разъединяем: {merged_range}")
            ws.unmerge_cells(str(merged_range))
    
    # Шаг 2: Удаляем первую строку
    print("\n>>> Удаление первой строки...")
    ws.delete_rows(1)
    
    # Шаг 3: Записываем правильные заголовки в новую первую строку (бывшую вторую)
    print("\n>>> Запись новых заголовков...")
    for col_idx, header in enumerate(CORRECT_HEADERS, start=1):
        ws.cell(row=1, column=col_idx).value = header
        print(f"  {chr(64+col_idx)}: {header}")
    
    # Сохраняем
    print("\n>>> Сохранение файла...")
    wb.save(str(file_path.absolute()))
    wb.close()
    
    # Проверяем результат
    print("\n>>> Проверка результата...")
    wb_check = load_workbook(file_path)
    ws_check = wb_check.active
    
    success = True
    for i in range(1, 17):
        value = ws_check.cell(1, i).value
        expected = CORRECT_HEADERS[i-1]
        if value != expected:
            print(f"  ✗ {chr(64+i)}: ожидалось '{expected}', получено '{value}'")
            success = False
    
    wb_check.close()
    
    if success:
        print("\n✅ Файл успешно обработан!")
    else:
        print("\n⚠️ Обнаружены расхождения!")
    
    return success

def main():
    """Обрабатывает все файлы в папке data/11.12.2025."""
    data_dir = Path("data/11.12.2025")
    
    if not data_dir.exists():
        print(f"Папка {data_dir} не найдена!")
        return
    
    # Получаем все xlsx файлы (исключая временные ~$)
    xlsx_files = [f for f in data_dir.glob("*.xlsx") if not f.name.startswith("~$")]
    
    if not xlsx_files:
        print(f"Файлы .xlsx не найдены в {data_dir}")
        return
    
    print(f"\n{'='*60}")
    print(f"Найдено файлов: {len(xlsx_files)}")
    print('='*60)
    
    results = {}
    for file_path in xlsx_files:
        try:
            success = fix_file_headers(file_path)
            results[file_path.name] = success
        except Exception as e:
            print(f"\n✗ Ошибка: {e}")
            import traceback
            traceback.print_exc()
            results[file_path.name] = False
    
    # Итоговый отчёт
    print(f"\n{'='*60}")
    print("ИТОГОВЫЙ ОТЧЁТ")
    print('='*60)
    for filename, success in results.items():
        status = "✅ OK" if success else "❌ ОШИБКА"
        print(f"{status}: {filename}")
    
    success_count = sum(1 for s in results.values() if s)
    total_count = len(results)
    
    print(f"\n{'='*60}")
    print(f"✅ Успешно обработано: {success_count}/{total_count}")
    print('='*60)

if __name__ == "__main__":
    main()








