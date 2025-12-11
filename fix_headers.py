"""Скрипт для исправления заголовков в файлах отчётов."""
from pathlib import Path
from openpyxl import load_workbook

# Правильные заголовки по столбцам A-P
CORRECT_HEADERS = [
    "Бренд",           # A
    "Предмет",         # B
    "Сезон",           # C
    "Коллекция",       # D
    "Наименование",    # E
    "Артикул поставщика",  # F
    "Номенклатура",    # G
    "Баркод",          # H
    "Размер",          # I
    "Контракт",        # J
    "Склад",           # K
    "Заказано шт",     # L
    "Заказано себестоимость",  # M
    "Выкупили шт",     # N
    "Выкупили руб",    # O
    "Текущий остаток"  # P
]

def fix_file_headers(file_path: Path):
    """Исправляет заголовки в файле.
    
    1. Удаляет первую строку (неполные заголовки)
    2. Заменяет новую первую строку на правильные заголовки
    """
    print(f"\nОбработка: {file_path.name}")
    
    # Загружаем файл
    wb = load_workbook(file_path)
    ws = wb.active
    
    # Показываем старую первую строку
    old_first_row = [ws.cell(1, i).value for i in range(1, 17)]
    print(f"  Старая первая строка: {old_first_row[:5]}...")
    
    # Удаляем первую строку
    ws.delete_rows(1)
    print("  ✓ Первая строка удалена")
    
    # Записываем правильные заголовки в новую первую строку
    for col_idx, header in enumerate(CORRECT_HEADERS, start=1):
        ws.cell(row=1, column=col_idx, value=header)
    
    # Проверяем новую первую строку
    new_first_row = [ws.cell(1, i).value for i in range(1, 17)]
    print(f"  Новая первая строка: {new_first_row[:5]}...")
    
    # Сохраняем
    wb.save(file_path)
    print(f"  ✓ Файл сохранён")

def main():
    """Обрабатывает все файлы в папке data/11.12.2025."""
    data_dir = Path("data/11.12.2025")
    
    if not data_dir.exists():
        print(f"Папка {data_dir} не найдена!")
        return
    
    # Получаем все xlsx файлы
    xlsx_files = list(data_dir.glob("*.xlsx"))
    
    if not xlsx_files:
        print(f"Файлы .xlsx не найдены в {data_dir}")
        return
    
    print(f"Найдено файлов: {len(xlsx_files)}")
    
    for file_path in xlsx_files:
        try:
            fix_file_headers(file_path)
        except Exception as e:
            print(f"  ✗ Ошибка: {e}")
    
    print("\n✅ Обработка завершена!")

if __name__ == "__main__":
    main()

