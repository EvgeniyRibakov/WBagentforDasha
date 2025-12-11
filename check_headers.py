"""Проверка заголовков в файле."""
from openpyxl import load_workbook

file_path = 'data/11.12.2025/beautylab_11.12.2025.xlsx'
wb = load_workbook(file_path)
ws = wb.active

print('Первая строка (заголовки):')
for i in range(1, 17):
    col_letter = chr(64 + i)  # A=65, B=66, etc.
    value = ws.cell(1, i).value
    print(f'{col_letter}: {value}')

print('\nВторая строка (первые 5 значений - данные):')
for i in range(1, 6):
    print(f'{ws.cell(2, i).value}', end=', ')
print()

